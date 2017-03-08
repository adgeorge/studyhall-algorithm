# Note:The template file will be copied to a new file. When you change the code of the template file you can create new file with this base code.
from random import randint
# from operator import attrgetter
from openpyxl import Workbook
from openpyxl.compat import range
# from openpyxl.cell import get_column_letter

def main():
# -----------------------------------Classes-----------------------------------------

    # Defines 'Student' Class
    class Student:
        def __init__(self, id):
            self.id = id
            self.study_group_list =[]
        
        def add_to_study_group_list(self, study_group, number_of_periods):
            if len(self.study_group_list) < number_of_periods:
                self.study_group_list.append(study_group)
            
        def print(self):
            print("student:", self.id, "is in study group(s):", self.study_group_list)

    # Defines 'Conflict' Class
    class Conflict:
        def __init__(self, conflicting_study_group, conflicting_list_of_students):
            self.conflicting_study_group = conflicting_study_group
            self.conflicting_list_of_students = conflicting_list_of_students
            self.number_of_conflicting_students = len(conflicting_list_of_students)

    # Defines 'Period' Class
    class Period:
        def __init__(self,id):
            self.id = id
            self.list_of_study_groups = []
            self.list_of_conflicting_students =[]
            self.number_of_conflicts = 0
             
        def find_conflicts(self):
            for current_study_group in range(0,len(self.list_of_study_groups)):
                for study_group_to_compare in range(0, len(self.list_of_study_groups)):
                    study_group1 = self.list_of_study_groups[current_study_group]
                    study_group2 = self.list_of_study_groups[study_group_to_compare]
                    if current_study_group < study_group_to_compare:
                        conflicting_students = list(set(study_group1.student_list) & set(study_group2.student_list))
                        if len(conflicting_students) > 0:
                            self.number_of_conflicts = self.number_of_conflicts + len(conflicting_students)
                            for student in conflicting_students:
                                self.list_of_conflicting_students.append(student)
        
        def print(self):
            print("Period", self.id,":")
            print("List of Study Groups:", [x.name for x in self.list_of_study_groups])
            print("List of Conflicting Students:", [x.id for x in self.list_of_conflicting_students])
            print("Number of Conflicts:", self.number_of_conflicts)
            print("\n")

    #Defines 'Scenario' Class
    class Scenario:
        def __init__(self, id, period_list):
            self.id = id
            self.period_list = period_list
            self.total_number_of_conflicts = 0

        def calculate_scenarios_total_conflicts(self):
            for period in self.period_list:
                self.total_number_of_conflicts = self.total_number_of_conflicts + period.number_of_conflicts
                
        def print(self):
            print("Scenario",self.id,":")
            print("Total Number of Conflicts:", self.total_number_of_conflicts)
         
    #Defines 'Study Group' Class
    class StudyGroup:
        def __init__(self, name):
            self.name = name
            self.student_list = []
            self.conflicting_study_group_list = []
            self.total_number_of_conflicts = 0
            self.period = 0
            self.impact_number = 0
            self.conflicts_in_period0 = 0
            self.conflicts_in_period1 = 0
            # self.number_of_conflicting_study_groups = len(self.conflicting_study_group_list)
        
        def add_student(self, student):
            self.student_list.append(student)
        
        def get_total_conflicts(self):
            self.total_number_of_conflicts = 0
            for study_group in self.conflicting_study_group_list:
                for student in study_group.conflicting_list_of_students:
                    self.total_number_of_conflicts +=1
            return self.total_number_of_conflicts
            
        def print(self):
            print(self.name)
            print("The students in", self.name, "are:",[x.id for x in self.student_list])
            print("Total number of students in", self.name, ":", len(self.student_list))

        def print_conflicts(self):
            print("Total number of conflicts in", self.name, ":", self.get_total_conflicts())
            # !!!!!!
            print("The conflicting study groups with", self.name, "are:\n",[x.conflicting_study_group.name for x in self.conflicting_study_group_list])
            for conflict_object in self.conflicting_study_group_list:
                print("Students in", self.name," that are also in ", conflict_object.conflicting_study_group.name,":")
                print([x.id for x in conflict_object.conflicting_list_of_students])
            print("\n") 


# -----------------------------------Functions-----------------------------------------

    def scenarios_generator(list_of_all_study_groups, number_of_periods):
        number_of_study_groups_in_last_period = 0
        scenario_list = []
        scenario_number = 0
        min_conflicts = 2000
        best_scenario_list = []
        number_of_total_study_groups = len(list_of_all_study_groups)

        def create_period_list(number_of_periods):
            listOfPeriods = []
            for num in range(0,number_of_periods):
                listOfPeriods.append(Period(num))
            return listOfPeriods

        def create_scenario(period_list,list_of_all_study_groups):
            for period in period_list:
                for study_group in list_of_all_study_groups:
                    if study_group.period == period.id:
                        period.list_of_study_groups.append(study_group)

                period.find_conflicts()
            return Scenario(scenario_number,period_list)

        # Will continue to run until all the study groups moved into the last period
        while number_of_study_groups_in_last_period != number_of_total_study_groups:
            number_of_study_groups_in_last_period = 0
            i = 0

            while (list_of_all_study_groups[i].period != 0) & (i < number_of_total_study_groups):
                list_of_all_study_groups[i].period = 0
                i+=1
            list_of_all_study_groups[i].period = 1

            period_list = create_period_list(number_of_periods)
            s = create_scenario(period_list,list_of_all_study_groups)
            s.calculate_scenarios_total_conflicts()
            scenario_list.append(s)

            #For Testing
            print("Scenario", scenario_number)
            print("Total Number of Conflicts:",s.total_number_of_conflicts)
            for period in period_list:
                print("Number of Conflicts in Period",period.id,":",period.number_of_conflicts)
                print("Period:",period.id,[x.name for x in period.list_of_study_groups])

            # Finds lowest amount of conflicts
            if s.total_number_of_conflicts < min_conflicts:
                min_conflicts = s.total_number_of_conflicts

            for study_group in range(0, number_of_total_study_groups):
                if list_of_all_study_groups[study_group].period == number_of_periods-1:
                    number_of_study_groups_in_last_period +=1

            #For Testing
            print("number of study groups in period 1:", number_of_study_groups_in_last_period)
            print("length of list of all study groups:", number_of_total_study_groups)
            print("\n")

            scenario_number +=1

        #Finds the Scenarios with the lowest amount of conflicts
        for scenario in scenario_list:
                if scenario.total_number_of_conflicts == min_conflicts:
                    best_scenario_list.append(scenario)
        print("Generated all scenarios!")
        print("Best Scenarios:", [x.id for x in best_scenario_list])
        print("With",min_conflicts,"number of conflicts")
        return scenario_list

    def generate_list_of_students(number_of_students_to_create):
        list_of_all_students = []
        for x in range(1, number_of_students_to_create + 1):
            list_of_all_students.append(Student(x))
        return list_of_all_students

    def generate_study_group_list(list_of_all_study_group_names):
        list_of_all_study_groups = []
        for x in list_of_all_study_group_names:
            list_of_all_study_groups.append(StudyGroup(x))
        return list_of_all_study_groups

    def assign_study_groups_to_students(list_of_students, number_of_periods,list_of_all_study_group_names):
        number_of_students = len(list_of_students)
        for x in range(0,number_of_students):
            study_group = 0
            while study_group < randint(0,number_of_periods):
                list_of_students[x].add_to_study_group_list(
                    list_of_all_study_group_names[
                        randint(0,len(list_of_all_study_group_names)-1)],number_of_periods)
                study_group += 1

    def assign_student_to_study_group(list_of_students, study_group_list):
        for current_study_group in study_group_list:
            for current_student in list_of_students:
                for index in current_student.study_group_list:
                    if index == current_study_group.name:
                        current_study_group.add_student(current_student)

    def get_study_group_conflicts(study_group_list):
        for current_study_group in range(0,len(study_group_list)):
            for study_group_to_compare in range(0, len(study_group_list)):
                study_group1 = study_group_list[current_study_group]
                study_group2 = study_group_list[study_group_to_compare]
                if current_study_group < study_group_to_compare:
                    conflicting_students = list(set(study_group1.student_list) & set(study_group2.student_list))
                    if len(conflicting_students) > 0:
                        # !!!!!
                        study_group1.conflicting_study_group_list.append(Conflict(study_group2, conflicting_students))
                        study_group2.conflicting_study_group_list.append(Conflict(study_group1, conflicting_students))
                        study_group1.total_number_of_conflicts = study_group1.total_number_of_conflicts + len(conflicting_students)
                        study_group2.total_number_of_conflicts = study_group1.total_number_of_conflicts + len(conflicting_students)

    def impact_number_algorithm(study_group_list, number_of_periods):
        study_group_list_ordered_by_impact_number = []

        def find_worst_pairs(study_group_list):
            worst_pairs_list = []

            def find_highest_number_of_conflicts(study_group_list):
                highest_conflict_number = 0
                for study_group in study_group_list:
                    for conflict in study_group.conflicting_study_group_list:
                        if conflict.number_of_conflicting_students > highest_conflict_number:
                            highest_conflict_number = conflict.number_of_conflicting_students
                return highest_conflict_number

            highest_conflict_number = find_highest_number_of_conflicts(study_group_list)

            for study_group in study_group_list:
                for conflict in study_group.conflicting_study_group_list:
                    if conflict.number_of_conflicting_students == highest_conflict_number:
                        worst_pairs_list.append(conflict.conflicting_study_group)
            return worst_pairs_list




# ------------------------------Functions for Testing--------------------------------

    def print_all_students_info(list_of_students):
        print("PRINTING STUDENTS")
        for student in list_of_students:
            student.print()
        print("FINISHED PRINTING STUDENTS")

    def print_all_study_group_info(study_group_list):
        print("STUDY GROUPS")
        for study_group in study_group_list:
            study_group.print()
            study_group.print_conflicts()
        print("FINISHED PRINTING STUDY GROUPS")

    def create_excel_file(scenario_list):
        wb = Workbook()

        ws = wb.active

        destination_filename = "UB Algorithm Data.xlsx"

        ws.sheet_properties.tabColor = "1072BA"

        a = ws.cell(row=1, column=1)
        b = ws.cell(row=1, column=2)
        a.value = "Scenario ID"
        b.value = "Number of Conflicts"

        ws.column_dimensions["B"].width = 16
        for x in range(0, len(scenario_list)):
            a = ws.cell(row=x + 2, column=1)
            a.value = scenario_list[x].id
            conflicts = ws.cell(row=x + 2, column=2)
            conflicts.value = scenario_list[x].total_number_of_conflicts

        wb.save(filename=destination_filename)

# -----------------------------------Main()-----------------------------------------
    # List of Study Group Names
    list_of_all_study_group_names = ["Integrated Math", "PreCalc", "Calculus",
                              "Stats & Probability", "Physics", "Chemistry",
                              "Anatomy", "Biology", "Environmental Science",
                              "English 10", "English 11", "English 12", "Vietnamese",
                              "Chinese", "Nepali", "STEM Lab", "Art", "Music",
                              "Senior Seminar"]


    # For testing
    # list_of_all_study_group_names = ["Integrated Math", "PreCalc", "Calculus",
    #                             "Stats & Probability", "Physics", "Chemistry",
    #                             "Anatomy", "Biology", "Environmental Science",
    #                             "English 10"]

    # This will determine how many periods are available
    number_of_periods = 2

    # Generate Students
    number_of_students_to_create = 200
    list_of_students = generate_list_of_students(number_of_students_to_create)

    # Randomly assigns study groups to students.
    # The number of study groups depends on the number of periods
    assign_study_groups_to_students(list_of_students,number_of_periods,list_of_all_study_group_names)

    # Creates a list of study groups based
    # on the array of study group names
    study_group_list = generate_study_group_list(list_of_all_study_group_names)

    # Matches up the students to their randomly assigned study groups
    assign_student_to_study_group(list_of_students, study_group_list)

    # Finds the number of students that are
    # assigned to more than one study group
    get_study_group_conflicts(study_group_list)

    # For testing purposes
    print_all_students_info(list_of_students)
    print_all_study_group_info(study_group_list)

    # Creates every possible scenario by placing every
    # combination of study groups in the given periods
# Uncomment this for comparing
    scenarios_list = scenarios_generator(study_group_list,number_of_periods)
    ordered_scenario_list = sorted(scenarios_list, key= lambda x: x.total_number_of_conflicts, reverse=False)

    # Create the EXCEL DOC
    # Sorted by total number of conflicts
# Uncomment this for comparing
    create_excel_file(ordered_scenario_list)


    # TEST ALGORITHM 1
    def algorithm1(study_group_list,number_of_periods):

        # How to assign the first two study groups
        def find_worst_pairs(study_group_list):
            # Iterates through a study_group_list and finds
            # study groups that have the highest number of
            # conflicts with each other. (Finds the study groups
            # that should not be in the same period.)
            max_conflict_number = 0
            worst_pairs_list = []
            for study_group in study_group_list:
                for conflict in study_group.conflicting_study_group_list:
                    if conflict.number_of_conflicting_students > max_conflict_number:
                        max_conflict_number = conflict.number_of_conflicting_students

            for study_group in study_group_list:
                for conflict in study_group.conflicting_study_group_list:
                    if conflict.number_of_conflicting_students == max_conflict_number:
                        worst_pairs_list.append(study_group)
            return worst_pairs_list

        def find_study_group_with_highest_total_conflict(study_group_list):
            print("\nTEST INFO: More than one worst pair exists")
            print("TEST INFO: Length of worst pairs list:",len(study_group_list))
            print([x.name for x in study_group_list])
            print("TEST INFO: Going to find the study group with the highest amount of total conflicts.")
            highest_total_conflicts = 0
            highest_total_conflict_study_group_list = []
            for study_group in study_group_list:
                if study_group.total_number_of_conflicts > highest_total_conflicts:
                    highest_total_conflicts = study_group.total_number_of_conflicts
            for study_group in study_group_list:
                if study_group.total_number_of_conflicts == highest_total_conflicts:
                    if study_group not in highest_total_conflict_study_group_list:
                        highest_total_conflict_study_group_list.append(study_group)
            print("TEST INFO: These are the study groups that have the highest conflict number and the most total conflicts:")
            print([x.name for x in highest_total_conflict_study_group_list])
            print("TEST INFO: Highest total conflict number is:",highest_total_conflicts)
            return highest_total_conflict_study_group_list

        def find_study_group_with_highest_number_of_conflicting_study_groups(highest_total_conflict_study_group_list):
            highest_number_of_conflicting_study_groups = 0
            highest_number_of_conflicting_study_groups_list = []
            for study_group in highest_total_conflict_study_group_list:
                if len(study_group.conflicting_study_group_list) > highest_number_of_conflicting_study_groups:
                    highest_number_of_conflicting_study_groups = len(study_group.conflicting_study_group_list)
            for study_group in highest_total_conflict_study_group_list:
                if len(study_group.conflicting_study_group_list) == highest_number_of_conflicting_study_groups:
                    highest_number_of_conflicting_study_groups_list.append(study_group)
            print("TEST INFO: Study groups with the highest number of conflicting study groups:")
            print([x.name for x in highest_number_of_conflicting_study_groups_list])
            return highest_number_of_conflicting_study_groups_list

        def find_first_study_group_to_assign(worst_pairs_list):

            if len(worst_pairs_list) > 2:
                # If more than 1 worst pair exists then find
                # the study group with the highest number of
                # total conflicts.
                highest_total_conflict_study_group_list = find_study_group_with_highest_total_conflict(worst_pairs_list)

                if len(highest_total_conflict_study_group_list) > 1:
                    # If more than one study group has the highest number
                    # of total conflicts, find the study group that also has the
                    # highest number of conflicting study groups.
                    highest_number_of_conflicting_study_groups_list = \
                        find_study_group_with_highest_number_of_conflicting_study_groups(
                            highest_total_conflict_study_group_list)

                    if len(highest_number_of_conflicting_study_groups_list) > 1:
                        print("TEST INFO: More than one study group is ideal to use first.")
                        print("TEST INFO: Going to assign the first study group in list by default.")
                        first_study_group_to_assign = highest_number_of_conflicting_study_groups_list[0]
                        print("First study group to assign is:",first_study_group_to_assign.name)

                    else:
                        print("TEST INFO: One study group is ideal to assign first.")
                        first_study_group_to_assign = highest_number_of_conflicting_study_groups_list[0]
                        print("First study group to assign is:",first_study_group_to_assign.name)

                else:
                    # There is only one study group with the highest number
                    # of total conflicts
                    first_study_group_to_assign =  highest_total_conflict_study_group_list[0]
                    print("First study group to assign is:",first_study_group_to_assign.name)

            else:
                # only one worst pair exists
                # put the worst pair in the period list
                print("Only one worst pair exists")
                print("Length of worst pairs list:",len(worst_pairs_list))
                print([x.name for x in worst_pairs_list])
                first_study_group_to_assign =  worst_pairs_list[0]
                print("First study group to assign is:",first_study_group_to_assign.name)
            return first_study_group_to_assign

        def find_second_study_group_to_assign(first_study_group_to_assign):
            max_conflict_number = 0
            worst_match_for_first_study_group = []
            for conflict in first_study_group_to_assign.conflicting_study_group_list:
                if conflict.number_of_conflicting_students > max_conflict_number:
                    max_conflict_number = conflict.number_of_conflicting_students

            for conflict in first_study_group_to_assign.conflicting_study_group_list:
                if conflict.number_of_conflicting_students == max_conflict_number:
                    worst_match_for_first_study_group.append(conflict.conflicting_study_group)

            if len(worst_match_for_first_study_group) > 1:
                # If there is more than one study group that is
                # a worst pair for study group one then find the
                # study group with the highest conflict number
                print("TEST 2")
                highest_total_conflict_study_group_list = \
                    find_study_group_with_highest_total_conflict(worst_match_for_first_study_group)
                if len(highest_total_conflict_study_group_list) > 1:
                    highest_number_of_conflicting_study_groups_list = \
                        find_study_group_with_highest_number_of_conflicting_study_groups(
                            highest_total_conflict_study_group_list)
                    if len(highest_number_of_conflicting_study_groups_list) > 1:
                        second_study_group_to_assign = highest_number_of_conflicting_study_groups_list[0]
                        print("The second study group to assign is:",second_study_group_to_assign.name)
                    else:
                        second_study_group_to_assign = highest_number_of_conflicting_study_groups_list[0]
                        print("The second study group to assign is:",second_study_group_to_assign.name)
                else:
                    second_study_group_to_assign = highest_total_conflict_study_group_list[0]
                    print("The second study group to assign is:",second_study_group_to_assign.name)
            else:
                print("There is only one possible choice for 2nd study group.")
                second_study_group_to_assign = worst_match_for_first_study_group[0]
                print("The second study group to assign is:",second_study_group_to_assign.name,"\n")
            return second_study_group_to_assign

        def calculate_impact_numbers(remaining_study_groups,period_list):
            for study_group_to_calculate in remaining_study_groups:
                #for period in period_list:
                conflicts_in_period0 = 0
                conflicts_in_period1 = 0
                # finds the conflicts each remaining study group has with
                # the study groups assigned in period 0
                for study_group_assigned in period_list[0].list_of_study_groups:
                    for conflict in study_group_to_calculate.conflicting_study_group_list:
                        if conflict.conflicting_study_group == study_group_assigned:
                            #print(conflict.conflicting_study_group.name,":",study_group_assigned.name,"-MATCHED")
                            conflicts_in_period0 += conflict.number_of_conflicting_students
                        #else:
                            #print(conflict.conflicting_study_group.name,":",study_group_assigned.name,"-NO MATCH")
                #print("\n")
                # finds the conflicts each remaining study group has with
                # the study groups assigned in period 1
                for study_group_assigned in period_list[1].list_of_study_groups:
                    for conflict in study_group_to_calculate.conflicting_study_group_list:
                        if conflict.conflicting_study_group == study_group_assigned:
                            #print(conflict.conflicting_study_group.name,":",study_group_assigned.name,"-MATCHED")
                            conflicts_in_period1 += conflict.number_of_conflicting_students
                        #else:
                            #print(conflict.conflicting_study_group.name,":",study_group_assigned.name,"-NO MATCH")
                # Takes the difference between conflicts in period 1 and 0
                # and assigns it the the study groups impact number
                study_group_to_calculate.conflicts_in_period0 = conflicts_in_period0
                study_group_to_calculate.conflicts_in_period1 = conflicts_in_period1
                study_group_to_calculate.impact_number = abs(conflicts_in_period1-conflicts_in_period0)
            print("\n")
            for study_group in remaining_study_groups:
                print(study_group.name,"impact number:",study_group.impact_number)
            return remaining_study_groups

        worst_pairs_list = find_worst_pairs(study_group_list)
        first_study_group_to_assign = find_first_study_group_to_assign(worst_pairs_list)
        second_study_group_to_assign = find_second_study_group_to_assign(first_study_group_to_assign)

    # Assigning the rest of the study groups
        # Create the number of Periods needed
        period_list_algo1 = []
        for x in range(0,number_of_periods):
            period_list_algo1.append(Period(x))

        # Add the first study groups to each period
        period_list_algo1[0].list_of_study_groups.append(first_study_group_to_assign)
        period_list_algo1[1].list_of_study_groups.append(second_study_group_to_assign)

        # Print for TESTING - Print period info
        for period in period_list_algo1:
            period.print()

        # Copy the study group list to use to calculate the impact number
        remaining_study_groups = list(study_group_list)

        # Remove the first study groups already assigned to a period
        # from the remaining study groups list
        remaining_study_groups.remove(first_study_group_to_assign)
        remaining_study_groups.remove(second_study_group_to_assign)

        # Print for TESTING - The names of the study groups in the remaining
        # study group list
        print([study_group.name for study_group in remaining_study_groups])

        # Calculating the impact number for all the study groups in the
        # remaining study group list
        #remaining_study_groups = calculate_impact_numbers(remaining_study_groups,period_list)

        # Assign the study group with the highest impact number to
        # the period with it has the least amount of conflicts with.
        # If there is a tie, pick the study group that has the highest
        # number of conflicts, then number of conflicting study groups,
        # default to just the first one in list if all the above criteria
        # are tied

        def assign_remaining_study_groups(remaining_study_groups,period_list):
            while len(remaining_study_groups) >= 1:
                remaining_study_groups = calculate_impact_numbers(remaining_study_groups,period_list)

                def find_study_groups_with_highest_impact_number(remaining_study_groups):
                    study_groups_with_highest_impact_number = []
                    highest_impact_number = max(study_group.impact_number for study_group in remaining_study_groups)
                    print("Highest impact number:",highest_impact_number)
                    for study_group in remaining_study_groups:
                        if study_group.impact_number == highest_impact_number:
                            study_groups_with_highest_impact_number.append(study_group)

                    for study_group in study_groups_with_highest_impact_number:
                        print(study_group.name,"conflict number:",study_group.total_number_of_conflicts)
                    return study_groups_with_highest_impact_number

                def tie_breaker1_conflict_number(study_groups_with_highest_impact_number):
                    study_groups_with_highest_conflict_number = []
                    highest_total_conflict_number = max(
                        study_group.total_number_of_conflicts for study_group in study_groups_with_highest_impact_number)
                    print("Tie Breaker 1 - Highest conflict number:",highest_total_conflict_number)
                    for study_group in study_groups_with_highest_impact_number:
                        if study_group.total_number_of_conflicts == highest_total_conflict_number:
                            study_groups_with_highest_conflict_number.append(study_group)

                    for study_group in study_groups_with_highest_conflict_number:
                        print(study_group.name,"number of conflicting study groups:",len(
                            study_group.conflicting_study_group_list))
                    return study_groups_with_highest_conflict_number

                def tie_breaker2_number_of_conflicting_study_groups(study_groups_with_highest_conflict_number):
                    study_groups_with_highest_number_of_conflicting_study_groups = []
                    highest_number_of_conflicting_study_groups = max(
                        len(study_group.conflicting_study_group_list)
                        for study_group in study_groups_with_highest_conflict_number)
                    print("Tie Breaker 2 - Highest number of conflicting study groups:",
                          highest_number_of_conflicting_study_groups)
                    for study_group in study_groups_with_highest_conflict_number:
                        if len(study_group.conflicting_study_group_list) == highest_number_of_conflicting_study_groups:
                            study_groups_with_highest_number_of_conflicting_study_groups.append(study_group)

                    for study_group in study_groups_with_highest_number_of_conflicting_study_groups:
                        print(study_group.name)
                    return study_groups_with_highest_number_of_conflicting_study_groups

                study_groups_with_highest_impact_number = \
                    find_study_groups_with_highest_impact_number(remaining_study_groups)

                if len(study_groups_with_highest_impact_number) > 1:
                    study_groups_with_highest_conflict_number = \
                        tie_breaker1_conflict_number(study_groups_with_highest_impact_number)

                    if len(study_groups_with_highest_conflict_number) > 1:
                        study_groups_with_highest_number_of_conflicting_study_groups = \
                            tie_breaker2_number_of_conflicting_study_groups(study_groups_with_highest_conflict_number)
                        if len(study_groups_with_highest_number_of_conflicting_study_groups) > 1:
                            study_group_to_assign_next = \
                                study_groups_with_highest_number_of_conflicting_study_groups[0]
                        else:
                            study_group_to_assign_next = \
                                study_groups_with_highest_number_of_conflicting_study_groups[0]
                    else:
                        study_group_to_assign_next = study_groups_with_highest_conflict_number[0]
                else:
                    study_group_to_assign_next = study_groups_with_highest_impact_number[0]

                print("Conflicts with Period 0:",study_group_to_assign_next.conflicts_in_period0)
                print("Conflicts with Period 1:",study_group_to_assign_next.conflicts_in_period1)
                if study_group_to_assign_next.conflicts_in_period1 > study_group_to_assign_next.conflicts_in_period0:
                    period_list[0].list_of_study_groups.append(study_group_to_assign_next)
                    print(study_group_to_assign_next.name,"assigned to Period 0")
                else:
                    period_list[1].list_of_study_groups.append(study_group_to_assign_next)
                    print(study_group_to_assign_next.name,"assigned to Period 1")


                remaining_study_groups.remove(study_group_to_assign_next)
                assign_remaining_study_groups(remaining_study_groups,period_list)


        assign_remaining_study_groups(remaining_study_groups,period_list_algo1)
        for period in period_list_algo1:
                period.find_conflicts()
                period.print()
        print("done!")

    algorithm1(study_group_list,number_of_periods)


if __name__=="__main__":
    main()



